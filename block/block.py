import hgtk
from flask import Blueprint, request, render_template, flash, redirect, url_for
from openpyxl import load_workbook, Workbook
from konlpy.tag import Okt
import difflib
import pymysql
import urllib.request
from bs4 import BeautifulSoup

def searchWord(word):
    #사전 검색 url
    url = "https://stdict.korean.go.kr/api/search.do"
    option = "?certkey_no=1112&key=A8476D8061FAC1B57C5BCE8DA4CDCF28&method=exact"
    query = "&q=" + urllib.parse.quote(word)
    url_query = url + option + query

    #Open API 검색 요청 개체 설정
    request = urllib.request.Request(url_query)

    #검색 요청 및 처리
    response = urllib.request.urlopen(request)
    rescode = response.getcode()
    if(rescode == 200):
        return response.read().decode('utf-8')
    else:
        return None


def wordExistCheck(comment):
    #검색 질의 요청
    res = searchWord(comment)
    xmlsoup = BeautifulSoup(res,'html.parser')
    items = xmlsoup.find_all('item')
    # 단어의 존재 여부 확인
    if len(items) == 0:
        #단어가 없는 경우
        return False
    else:
        #단어가 있는 경우
     return True


def tokenize(comment):
    print("**품사 분리 시작**")
    okt=Okt()
    return [t for t in okt.nouns(comment)]
# 품사분리함수

def StringMatch(comment):

    load_wb = load_workbook("C:/Users/JAELYANG/Desktop/ICO/basic_keyword/공용keyword-3.xlsx", data_only=True)
    load_ws = load_wb['Sheet1']
    block = 0
    _comment = ""
    print("**1차 필터링 시작**")

    _comment = onlyHangul(comment)

    for i in range(1, 1103):
        # 차단 키워드 갯수만큼 for문
        if _comment.find(str(load_ws['A' + str(i)].value)) != -1:
            block = block + 1
            print("매치된 기본 키워드: " + load_ws['A' + str(i)].value)
            break
    #    한글 이외의 것을 제거한 댓글과 키워드 매치
    if block != 0:
        return load_ws['A' + str(i)].value
    else:
        return "OK"
#String 일치함수, 1차필터링

def onlyHangul(comment):
#특수문자 제거 함수
    _comment = ""
    for j in range(0, len(comment)):
        # 댓글 길이만큼 for문
        if hgtk.checker.is_hangul(comment[j]) :
            _comment += comment[j]
        elif comment[j] == ' ':
            _comment += comment[j]
        #     코멘트 한글자마다 한글인지 파악
        #     한글일 경우 새 String인자에 추가
        else:
            continue
        #      한글이 아니면 추가X
    return _comment
#띄어쓰기,특수문자 제외 한글만 추출하는 함수

def filteringSynk(comment):
    _comment = ""
    load_wb = load_workbook("C:/Users/JAELYANG/Desktop/ICO/basic_keyword/기본키워드_분리4.xlsx", data_only=True)
    load_ws = load_wb['Sheet']
    block = 0

    print("**2차 필터링 시작**")
    for j in comment:
        _comment = hgtk.text.decompose(j).replace("ᴥ", "")

        for i in range(1,824):
            matchRatio = difflib.SequenceMatcher(None,load_ws['A' + str(i)].value, _comment).ratio()

            if matchRatio >= 0.75:
                # 일치도 75%이상일시 단어가 국어사전에존재하는지 여부 확인, 존재하면 욕X,아니면 욕
                if wordExistCheck(j):
                    print("\t 존재하는 단어 :" + j + "이므로 차단하지 않습니다")
                    continue
                else:
                    print("기본 키워드: " + load_ws['A' + str(i)].value)
                    print("댓글 내 단어: " + _comment)
                    print("일치율: " + str(matchRatio*100) + "%")
                    block = block + 1
                    break
        if block !=0:
            break
    if block !=0:
        return load_ws['A' + str(i)].value + " & " + _comment + str(matchRatio*100) + "%"
    else:
        return "OK"
#유사도판별함수, 2차필터링


def runBlockComment(testComment):

        blokcedComment = "차단되었습니다."

        filtering1 = StringMatch(testComment)
        # 1차 필터링~String일치로 판별

        if filtering1 == "OK":
            # 1차 성공시 2차 필터링 시작
            testTokenComment = tokenize(testComment)
            # 품사분리(명사만 추출)
            filtering2 = filteringSynk(testTokenComment)
            # 자모음 분리 후 2차 필터링

            if filtering2 == "OK":
                return testComment
            ################################# 2차도 OK면 클린, ML로 댓글 넘기는 부분 필요합니다
            else:
                return blokcedComment
            # 2차에서 걸린경우
        else:
            return blokcedComment
            # 1차에서 걸린경우
